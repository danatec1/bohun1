"""
Folium Map Service
Folium을 사용한 지도 생성 서비스
"""

import folium
import os
from datetime import datetime
from typing import List, Dict, Any
from ..repositories.testdb_hospital_repository import TestDBHospitalRepository


class FoliumMapService:
    def __init__(self):
        self.repository = TestDBHospitalRepository()
        
    def create_hospital_map(self, 
                          center_lat: float = 36.5, 
                          center_lng: float = 127.5, 
                          zoom_start: int = 7) -> str:
        """
        병원 위치를 표시한 Folium 지도 생성
        
        Args:
            center_lat: 지도 중심 위도 (기본: 한국 중심)
            center_lng: 지도 중심 경도 (기본: 한국 중심)
            zoom_start: 초기 줌 레벨
            
        Returns:
            생성된 HTML 파일의 경로
        """
        
        # 병원 데이터 가져오기
        hospitals = self.repository.find_all()
        
        # Folium 지도 생성 - OpenStreetMap을 기본 타일로 설정
        m = folium.Map(
            location=[center_lat, center_lng],
            zoom_start=zoom_start,
            tiles='OpenStreetMap',
            attr='OpenStreetMap',
            prefer_canvas=True
        )
        
        # 다양한 타일 레이어 추가 (OpenStreetMap이 기본으로 체크됨)
        # 1. VWorld 기본지도 (한글)
        folium.TileLayer(
            tiles='https://map.vworld.kr/tile/Base/{z}/{x}/{y}.png',
            attr='VWorld 한국지도',
            name='VWorld 기본지도 (한글)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 2. VWorld 위성지도 (한글 지명)
        folium.TileLayer(
            tiles='https://map.vworld.kr/tile/Satellite/{z}/{x}/{y}.jpeg',
            attr='VWorld 위성지도',
            name='VWorld 위성지도 (한글)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 3. VWorld 하이브리드 (위성+지명 한글)
        folium.TileLayer(
            tiles='https://map.vworld.kr/tile/Hybrid/{z}/{x}/{y}.png',
            attr='VWorld 하이브리드',
            name='VWorld 하이브리드 (한글)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 4. CartoDB Positron (영문 - 밝은 테마)
        folium.TileLayer(
            tiles='CartoDB positron',
            name='밝은 지도 (영문)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 5. CartoDB Dark (영문 - 어두운 테마)
        folium.TileLayer(
            tiles='CartoDB dark_matter',
            name='어두운 지도 (영문)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 종별 FeatureGroup 생성 (체크박스로 ON/OFF 가능)
        feature_groups = {
            '종합병원': folium.FeatureGroup(name='종합병원', show=True),
            '병원': folium.FeatureGroup(name='병원', show=True),
            '의원': folium.FeatureGroup(name='의원', show=True),
            '요양병원': folium.FeatureGroup(name='요양병원', show=True)
        }
        
        # 병원 마커 추가
        hospital_count = 0
        type_stats = {
            '종합병원': 0,
            '병원': 0,
            '의원': 0,
            '요양병원': 0,
            '기타': 0
        }
        
        for hospital in hospitals:
            if hospital.latitude and hospital.longitude:
                # 팝업 내용 생성
                popup_html = self._create_popup_html(hospital)
                
                # testdb.위탁병원현황의 종별 컬럼 사용
                hospital_type = hospital.hospital_type if hasattr(hospital, 'hospital_type') else '기타'
                
                # 종별 통계 업데이트
                if hospital_type in type_stats:
                    type_stats[hospital_type] += 1
                else:
                    type_stats['기타'] += 1
                
                # 종별에 따른 마커 색상 결정
                marker_color = self._get_type_marker_color(hospital_type)
                marker_icon = self._get_type_marker_icon(hospital_type)
                icon_color = 'white'
                
                # 마커 생성
                marker = folium.Marker(
                    location=[hospital.latitude, hospital.longitude],
                    popup=folium.Popup(popup_html, max_width=300),
                    tooltip=f"{hospital.name} ({hospital_type})",
                    icon=folium.Icon(
                        color=marker_color,
                        icon=marker_icon,
                        prefix='fa',
                        icon_color=icon_color
                    )
                )
                
                # 해당 종별의 FeatureGroup에 마커 추가
                if hospital_type in feature_groups:
                    marker.add_to(feature_groups[hospital_type])
                else:
                    marker.add_to(m)  # 기타는 기본 지도에 추가
                
                hospital_count += 1
        
        # 모든 FeatureGroup을 지도에 추가
        for group in feature_groups.values():
            group.add_to(m)
        
        # 레이어 컨트롤 추가 (종별 체크박스 포함) - 오른쪽 상단
        folium.LayerControl(position='topright', collapsed=False).add_to(m)
        
        # 정보 패널 추가 (종별 통계 포함)
        info_html = f"""
        <div style="position: fixed; 
                    top: 10px; left: 50px; width: 280px; 
                    background-color: white; border:2px solid grey; z-index:9999; 
                    font-size:13px; padding: 12px; border-radius: 8px;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.2);">
        <h4 style="margin: 0 0 10px 0;">🏥 위탁병원 지도</h4>
        <p style="margin: 5px 0;"><strong>총 병원 수:</strong> {hospital_count}개</p>
        <hr style="margin: 8px 0; border: none; border-top: 1px solid #ddd;">
        <p style="margin: 5px 0; font-weight: bold;">📍 종별 현황:</p>
        <p style="margin: 3px 0;">
            <span style="color: #d63333; font-weight: bold;">●</span> 종합병원: {type_stats['종합병원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #3498db; font-weight: bold;">●</span> 병원: {type_stats['병원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #38a169; font-weight: bold;">●</span> 의원: {type_stats['의원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #ff8c00; font-weight: bold;">●</span> 요양병원: {type_stats['요양병원']}개
        </p>
        <hr style="margin: 8px 0; border: none; border-top: 1px solid #ddd;">
        <p style="margin: 5px 0; font-size: 11px; color: #666;">
        💡 좌측 상단 체크박스로 종별 표시/숨김</p>
        <p style="margin: 5px 0; font-size: 11px; color: #666;">
        생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        </div>
        """
        m.get_root().html.add_child(folium.Element(info_html))
        
        # HTML 파일로 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'hospital_map_folium_{timestamp}.html'
        filepath = os.path.join(os.getcwd(), filename)
        
        m.save(filepath)
        
        return filepath
    
    def _create_popup_html(self, hospital) -> str:
        """병원 정보 팝업 HTML 생성"""
        # 병상수와 진료과수 포맷팅
        bed_info = f"{hospital.bed_count}병상" if hospital.bed_count else "-"
        dept_info = f"{hospital.department_count}과" if hospital.department_count else "-"
        
        popup_html = f"""
        <div style="width: 280px;">
            <h4 style="color: #2E86AB; margin-bottom: 10px;">🏥 {hospital.name}</h4>
            <p><strong>📍 주소:</strong><br>{hospital.address}</p>
            <p><strong>🏥 규모:</strong><br>
               병상수: {bed_info} | 진료과수: {dept_info}
            </p>
            <p><strong>🌍 좌표:</strong><br>
               위도: {hospital.latitude:.6f}<br>
               경도: {hospital.longitude:.6f}
            </p>
        </div>
        """
        return popup_html
    
    def create_region_map(self, region: str = None) -> str:
        """
        특정 지역의 병원 지도 생성
        
        Args:
            region: 지역명 (예: '서울', '부산' 등)
            
        Returns:
            생성된 HTML 파일의 경로
        """
        hospitals = self.repository.find_all()
        
        # 지역별 필터링
        if region:
            filtered_hospitals = [h for h in hospitals if region in h.address]
        else:
            filtered_hospitals = hospitals
        
        if not filtered_hospitals:
            return None
        
        # 지역별 중심 좌표 및 줌 레벨 설정
        region_centers = {
            '서울': (37.5665, 126.9780, 11),
            '부산': (35.1796, 129.0756, 11),
            '대구': (35.8714, 128.6014, 11),
            '인천': (37.4563, 126.7052, 11),
            '광주': (35.1595, 126.8526, 11),
            '대전': (36.3504, 127.3845, 11),
            '울산': (35.5384, 129.3114, 11),
            '세종': (36.4800, 127.2890, 11),
            '경기': (37.4138, 127.5183, 9),
            '강원': (37.8228, 128.1555, 9),
            '충북': (36.8000, 127.7000, 9),
            '충남': (36.5184, 126.8000, 9),
            '전북': (35.7175, 127.1530, 9),
            '전남': (34.8679, 126.9910, 9),
            '경북': (36.4919, 128.8889, 9),
            '경남': (35.4606, 128.2132, 9),
            '제주': (33.4890, 126.4983, 10)
        }
        
        # 지역에 맞는 중심 좌표와 줌 레벨 가져오기
        if region and region in region_centers:
            center_lat, center_lng, zoom_level = region_centers[region]
        else:
            # 필터링된 병원들의 평균 위치 계산
            valid_hospitals = [h for h in filtered_hospitals 
                             if h.latitude and h.longitude]
            if valid_hospitals:
                center_lat = sum(h.latitude for h in valid_hospitals) / len(valid_hospitals)
                center_lng = sum(h.longitude for h in valid_hospitals) / len(valid_hospitals)
                zoom_level = 10
            else:
                center_lat, center_lng, zoom_level = 36.5, 127.5, 7
        
        # 지도 생성
        m = folium.Map(
            location=[center_lat, center_lng],
            zoom_start=zoom_level,
            tiles='OpenStreetMap',
            attr='OpenStreetMap',
            prefer_canvas=True
        )
        
        # 다양한 타일 레이어 추가
        folium.TileLayer(
            tiles='https://map.vworld.kr/tile/Base/{z}/{x}/{y}.png',
            attr='VWorld 한국지도',
            name='VWorld 기본지도 (한글)',
            overlay=False,
            control=True,
            show=False
        ).add_to(m)
        
        # 종별 FeatureGroup 생성
        feature_groups = {
            '종합병원': folium.FeatureGroup(name='종합병원', show=True),
            '병원': folium.FeatureGroup(name='병원', show=True),
            '의원': folium.FeatureGroup(name='의원', show=True),
            '요양병원': folium.FeatureGroup(name='요양병원', show=True)
        }
        
        # 병원 마커 추가 및 통계
        hospital_count = 0
        type_stats = {
            '종합병원': 0,
            '병원': 0,
            '의원': 0,
            '요양병원': 0,
            '기타': 0
        }
        
        for hospital in filtered_hospitals:
            if hospital.latitude and hospital.longitude:
                popup_html = self._create_popup_html(hospital)
                
                # 종별 확인
                hospital_type = hospital.hospital_type if hasattr(hospital, 'hospital_type') else '기타'
                
                # 통계 업데이트
                if hospital_type in type_stats:
                    type_stats[hospital_type] += 1
                else:
                    type_stats['기타'] += 1
                
                # 마커 색상 및 아이콘
                marker_color = self._get_type_marker_color(hospital_type)
                marker_icon = self._get_type_marker_icon(hospital_type)
                
                marker = folium.Marker(
                    location=[hospital.latitude, hospital.longitude],
                    popup=folium.Popup(popup_html, max_width=300),
                    tooltip=f"{hospital.name} ({hospital_type})",
                    icon=folium.Icon(
                        color=marker_color,
                        icon=marker_icon,
                        prefix='fa',
                        icon_color='white'
                    )
                )
                
                # FeatureGroup에 추가
                if hospital_type in feature_groups:
                    marker.add_to(feature_groups[hospital_type])
                else:
                    marker.add_to(m)
                
                hospital_count += 1
        
        # FeatureGroup을 지도에 추가
        for group in feature_groups.values():
            group.add_to(m)
        
        # 레이어 컨트롤 추가
        folium.LayerControl(position='topright', collapsed=False).add_to(m)
        
        # 정보 패널 추가
        region_text = f" - {region}" if region else ""
        info_html = f"""
        <div style="position: fixed; 
                    top: 10px; left: 50px; width: 280px; 
                    background-color: white; border:2px solid grey; z-index:9999; 
                    font-size:13px; padding: 12px; border-radius: 8px;
                    box-shadow: 0 2px 8px rgba(0,0,0,0.2);">
        <h4 style="margin: 0 0 10px 0;">🏥 위탁병원 지도{region_text}</h4>
        <p style="margin: 5px 0;"><strong>총 병원 수:</strong> {hospital_count}개</p>
        <hr style="margin: 8px 0; border: none; border-top: 1px solid #ddd;">
        <p style="margin: 5px 0; font-weight: bold;">📍 종별 현황:</p>
        <p style="margin: 3px 0;">
            <span style="color: #d63333; font-weight: bold;">●</span> 종합병원: {type_stats['종합병원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #3498db; font-weight: bold;">●</span> 병원: {type_stats['병원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #38a169; font-weight: bold;">●</span> 의원: {type_stats['의원']}개
        </p>
        <p style="margin: 3px 0;">
            <span style="color: #ff8c00; font-weight: bold;">●</span> 요양병원: {type_stats['요양병원']}개
        </p>
        <hr style="margin: 8px 0; border: none; border-top: 1px solid #ddd;">
        <p style="margin: 5px 0; font-size: 11px; color: #666;">
        생성: {datetime.now().strftime('%Y-%m-%d %H:%M')}</p>
        </div>
        """
        m.get_root().html.add_child(folium.Element(info_html))
        
        # 파일 저장
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        region_name = region.replace(' ', '_') if region else 'all'
        filename = f'hospital_map_{region_name}_{timestamp}.html'
        filepath = os.path.join(os.getcwd(), filename)
        
        m.save(filepath)
        
        return filepath
    
    def export_to_excel(self, filename: str = None) -> str:
        """
        병원 데이터를 Excel 파일로 내보내기 (openpyxl 사용)
        
        Args:
            filename: 저장할 파일명
            
        Returns:
            생성된 Excel 파일의 경로
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils.dataframe import dataframe_to_rows
        import pandas as pd
        
        # 병원 데이터 가져오기
        hospitals = self.repository.find_all()
        
        # 데이터를 딕셔너리 리스트로 변환
        data = []
        for hospital in hospitals:
            data.append({
                'ID': hospital.hospital_id,
                '병원명': hospital.name,
                '주소': hospital.address,
                '위도': hospital.latitude,
                '경도': hospital.longitude,
                '진료과목': ', '.join(hospital.medical_departments) if hospital.medical_departments else ''
            })
        
        # pandas DataFrame 생성
        df = pd.DataFrame(data)
        
        # 워크북 생성
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "병원현황"
        
        # 헤더 스타일
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # 데이터 추가
        for r in dataframe_to_rows(df, index=False, header=True):
            ws.append(r)
        
        # 헤더 스타일 적용
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 컬럼 너비 자동 조정
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # 파일 저장
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'hospital_data_{timestamp}.xlsx'
        
        filepath = os.path.join(os.getcwd(), filename)
        wb.save(filepath)
        
        return filepath
    
    def _get_hospital_type(self, medical_departments) -> str:
        """병원 유형 결정"""
        if not medical_departments:
            return '일반병원'
        
        departments_str = ' '.join(medical_departments)
        
        if '종합병원' in departments_str:
            return '종합병원'
        elif '재활' in departments_str:
            return '재활병원'
        elif '요양' in departments_str:
            return '요양병원'
        elif '한방' in departments_str or '한의' in departments_str:
            return '한방병원'
        else:
            return '일반병원'
    
    def _get_marker_color(self, medical_departments) -> str:
        """병원 유형에 따른 마커 색상 결정"""
        hospital_type = self._get_hospital_type(medical_departments)
        
        color_map = {
            '종합병원': 'red',          # 진한 빨간색
            '재활병원': 'orange',       # 주황색
            '요양병원': 'green',        # 진한 초록색
            '한방병원': 'purple',       # 보라색
            '일반병원': 'blue'          # 파란색
        }
        
        return color_map.get(hospital_type, 'blue')
    
    def _get_marker_icon(self, medical_departments) -> str:
        """병원 유형에 따른 마커 아이콘 결정"""
        hospital_type = self._get_hospital_type(medical_departments)
        
        icon_map = {
            '종합병원': 'hospital-o',   # 종합병원
            '재활병원': 'wheelchair',   # 재활
            '요양병원': 'heartbeat',    # 요양
            '한방병원': 'leaf',         # 한방
            '일반병원': 'plus'          # 일반
        }
        
        return icon_map.get(hospital_type, 'plus')
    
    def _get_custom_icon_color(self, hospital_type: str) -> str:
        """병원 유형에 따른 아이콘 내부 색상 결정"""
        icon_color_map = {
            '종합병원': 'white',      # 빨간 배경에 흰색 아이콘
            '재활병원': 'white',      # 주황 배경에 흰색 아이콘
            '요양병원': 'white',      # 초록 배경에 흰색 아이콘
            '한방병원': 'white',      # 보라 배경에 흰색 아이콘
            '일반병원': 'white'       # 파랑 배경에 흰색 아이콘
        }
        return icon_color_map.get(hospital_type, 'white')
    
    def _get_type_marker_color(self, hospital_type: str) -> str:
        """종별에 따른 마커 색상 결정 (testdb.위탁병원현황 종별 기준)"""
        color_map = {
            '종합병원': 'red',        # 빨강
            '병원': 'blue',           # 파랑
            '의원': 'green',          # 초록
            '요양병원': 'orange'      # 주황
        }
        return color_map.get(hospital_type, 'gray')
    
    def _get_type_marker_icon(self, hospital_type: str) -> str:
        """종별에 따른 마커 아이콘 결정 (testdb.위탁병원현황 종별 기준)"""
        icon_map = {
            '종합병원': 'hospital-o',  # 종합병원
            '병원': 'plus',            # 병원
            '의원': 'user-md',         # 의원
            '요양병원': 'heartbeat'    # 요양병원
        }
        return icon_map.get(hospital_type, 'question')
